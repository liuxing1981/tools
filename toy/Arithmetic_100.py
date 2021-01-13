import time
start = time.time()
import os
import sys
import tkinter as tk
import random
from tkinter import messagebox
end = time.time()
print(end - start)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, colors
from openpyxl.utils import get_column_letter

def cal(limit=20, count=100):
    result = set()
    while len(result) < count:
        i = random.randint(1, limit)
        j = random.randint(1, limit)
        if i ==j:
            continue
        if add.get() == 1 and i + j <= limit:
            # 无进位
            if add_1.get() == 1:
                if i%10 + j%10 < 10:
                    result.add('%s + %s = ' % (i, j))
            else:
                result.add('%s + %s = ' % (i, j))
        if sub.get() == 1 and i -j <=limit:
            # 无退位
            if sub_1.get() == 1 and abs(i%10 - j%10) < 10:
                if i >= j:
                    result.add('%s - %s = '%(i,j))
                else:
                    result.add('%s - %s = ' % (j, i))
            else:
                if i >= j:
                    result.add('%s - %s = ' % (i, j))
                else:
                    result.add('%s - %s = ' % (j, i))
        if mul.get == 1 and i * j <=limit:
            result.add('%s × %s = ' % (i, j))
        if div.get() == 1 and i % j == 0 and i / j <= limit:
            if j == 1 or i == 1:
                continue
            if i >= j:
                result.add('%s ÷ %s = ' % (i, j))
            else:
                result.add('%s ÷ %s = ' % (j, i))
    return list(result)


def create():
    global page_value, limit_value, count_value
    page, limit, count = page_value.get(),limit_value.get(), count_value.get()
    font = Font('宋体', size = 14, bold = False, italic = False, strike = False, color = colors.BLACK)
    align = Alignment(horizontal='left', vertical ='center', wrap_text = True)
    wb = Workbook()
    for i in range(page):
        result = cal(limit, count)
        if i == 0:
            ws = wb.active
            ws.title = 'page1'
        else:
            ws = wb.create_sheet(index=i, title='page'+str(i+1))
        nameCell = ws.cell(row=1, column=4, value='姓名：')
        classCell = ws.cell(row=1, column=5, value='班级：')
        nameCell.font = font
        classCell.font = font
        perRow = 5
        perCol = count // perRow
        splitLine = len(result)//perRow + 3 - len(result)//2//perRow
        for i in range(3, perCol + 3 + 1):
            for j in range(1,perRow + 1):
                # insert a empty line for split to half
                if i == splitLine:
                    ws.cell(row=i, column=j)
                    continue
                if len(result) > 0:
                    content = random.choice(result)
                    myCell = ws.cell(row=i, column=j, value=content)
                    myCell.font = font
                    myCell.alignment = align
                    # print('write to %s,%s %s'%(i,j,content))
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = 30
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16
    wb.save('data.xlsx')
    messagebox.showinfo(title='success', message='题目已经生成，位于：' + os.path.join(os.getcwd(),'data.xlsx'))
    sys.exit(0)

start = time.time()
window = tk.Tk()
window.title('小学数学运算题目生成器')
window.resizable(0, 0)
screenWidth = window.winfo_screenwidth()  # 获取显示区域的宽度
screenHeight = window.winfo_screenheight()  # 获取显示区域的高度
width = 450  # 设定窗口宽度
height = 200  # 设定窗口高度
left = (screenWidth - width) / 2
top = (screenHeight - height) / 2
window.geometry("%dx%d+%d+%d" % (width, height, left, top))
limit_value = tk.IntVar()
limit_value.set(20)
limit_text = tk.Label(window, text='数字范围：')
limit = tk.Entry(window, show=None, font=('Arial', 12),width=3, textvariable=limit_value)

count_value = tk.IntVar()
count_value.set(100)
count_text = tk.Label(window, text='题目数量：')
count = tk.Entry(window, show=None, font=('Arial', 12),width=3, textvariable=count_value)
opt_text = tk.Label(window, text='运算符号：')
page_value = tk.IntVar()
page_value.set(5)
page_text = tk.Label(window, text='生成几份：')
page = tk.Entry(window, show=None, font=('Arial', 12),width=3, textvariable=page_value)
# check box
add = tk.IntVar()
sub = tk.IntVar()
add_1 = tk.IntVar()
sub_1 = tk.IntVar()
mul = tk.IntVar()
div = tk.IntVar()
add.set(1)
sub.set(1)
c1 = tk.Checkbutton(window, text='加法', variable=add, onvalue=1, offvalue=0)
c2 = tk.Checkbutton(window, text='减法', variable=sub, onvalue=1, offvalue=0)
c3 = tk.Checkbutton(window, text='无进位', variable=add_1, onvalue=1, offvalue=0)
c4 = tk.Checkbutton(window, text='无退位', variable=sub_1, onvalue=1, offvalue=0)
c5 = tk.Checkbutton(window, text='乘法', variable=mul, onvalue=1, offvalue=0)
c6 = tk.Checkbutton(window, text='除法', variable=div, onvalue=1, offvalue=0)
gen_button = tk.Button(window, text="生成题目", command=create)
limit_text.place(x=30, y=25)
limit.place(x=100, y=25)
count_text.place(x=30, y=60)
count.place(x=100, y=60)
opt_text.place(x=30, y=95)
c1.place(x=100, y=95)
c2.place(x=150, y=95)
c3.place(x=200, y=95)
c4.place(x=270, y=95)
c5.place(x=330, y=95)
c6.place(x=380, y=95)
page_text.place(x=30, y=130)
page.place(x=100, y=130)
gen_button.place(x=30, y=165)
end = time.time()
print(end - start)
window.mainloop()
